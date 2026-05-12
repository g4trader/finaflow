"""
API de Onboarding para novas empresas
Permite importação de dados de planilha Excel/Google Sheets em etapas
"""

from fastapi import APIRouter, HTTPException, Depends, Body, BackgroundTasks
from fastapi.responses import JSONResponse
from pydantic import BaseModel, HttpUrl, Field
from typing import Optional, Dict, Any, List
from datetime import datetime
from decimal import Decimal
import os
import sys
import re
import uuid
from pathlib import Path
import requests
import tempfile
import json
import logging
import time

from app.services.dependencies import get_current_active_user, require_super_admin
from app.models.auth import User, Tenant, BusinessUnit, AuditLog
from app.database import SessionLocal
from sqlalchemy.orm import Session
from sqlalchemy import and_

router = APIRouter(prefix="/api/v1/onboarding", tags=["onboarding"])
logger = logging.getLogger(__name__)

# Adicionar backend ao path para importar scripts (apenas quando necessário)
backend_path = Path(__file__).parent.parent.parent
# Não adicionar ao sys.path no nível do módulo - fazer apenas quando necessário

MONTH_LABELS = [
    "JANEIRO",
    "FEVEREIRO",
    "MARÇO",
    "ABRIL",
    "MAIO",
    "JUNHO",
    "JULHO",
    "AGOSTO",
    "SETEMBRO",
    "OUTUBRO",
    "NOVEMBRO",
    "DEZEMBRO",
]

class SpreadsheetUrlRequest(BaseModel):
    url: HttpUrl
    tenant_id: Optional[str] = None
    business_unit_id: Optional[str] = None

class ImportRequest(BaseModel):
    tenant_id: str
    business_unit_id: str
    spreadsheet_url: HttpUrl
    reset_data: bool = False
    corrections: Optional[Dict[str, Any]] = None

class StartOnboardingRequest(BaseModel):
    tenant_name: str
    tenant_cnpj: str
    business_unit_name: str
    business_unit_code: Optional[str] = None
    spreadsheet_url: HttpUrl
    reset_data: bool = False
    corrections: Optional[Dict[str, Any]] = None

class ClearDataRequest(BaseModel):
    tenant_id: str
    business_unit_id: str
    year: Optional[int] = None

class OnboardingStatus(BaseModel):
    tenant_id: str
    business_unit_id: str
    status: str  # not_started, validating, importing_plan, importing_transactions, reconciling, completed, error
    current_step: Optional[str] = None
    progress: int = 0  # 0-100
    message: Optional[str] = None
    errors: List[str] = Field(default_factory=list)
    stats: Optional[Dict[str, Any]] = None
    run_id: Optional[str] = None
    started_at: Optional[str] = None
    finished_at: Optional[str] = None
    last_updated_at: Optional[str] = None

# Armazenamento temporário do status (em produção, usar Redis ou banco)
onboarding_status: Dict[str, OnboardingStatus] = {}


def log_onboarding(event: str, **context: Any) -> None:
    payload = {"event": event, **context}
    logger.info("onboarding_event=%s", json.dumps(payload, ensure_ascii=True))


def init_status(tenant_id: str, business_unit_id: str, message: str, run_id: str) -> OnboardingStatus:
    now = datetime.utcnow().isoformat()
    return OnboardingStatus(
        tenant_id=str(tenant_id),
        business_unit_id=str(business_unit_id),
        status="validating",
        current_step="Validando planilha",
        progress=0,
        message=message,
        run_id=run_id,
        started_at=now,
        last_updated_at=now,
    )


def update_status(
    status_key: str,
    *,
    status: Optional[str] = None,
    current_step: Optional[str] = None,
    progress: Optional[int] = None,
    message: Optional[str] = None,
    stats: Optional[Dict[str, Any]] = None,
    error: Optional[str] = None,
    finished: bool = False,
) -> None:
    status_obj = onboarding_status.get(status_key)
    if not status_obj:
        return

    if status is not None:
        status_obj.status = status
    if current_step is not None:
        status_obj.current_step = current_step
    if progress is not None:
        status_obj.progress = progress
    if message is not None:
        status_obj.message = message
    if stats is not None:
        status_obj.stats = stats
    if error:
        status_obj.errors.append(error)
    status_obj.last_updated_at = datetime.utcnow().isoformat()
    if finished:
        status_obj.finished_at = status_obj.last_updated_at

def normalize_spreadsheet_url(spreadsheet_url: str) -> str:
    """Normaliza URL do Google Sheets para download em XLSX."""
    if "docs.google.com/spreadsheets" not in spreadsheet_url:
        return spreadsheet_url

    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", spreadsheet_url)
    if match:
        sheet_id = match.group(1)
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=xlsx"

    base_url = spreadsheet_url.split("?")[0].split("#")[0]
    if "/edit" in base_url:
        base_url = base_url.split("/edit")[0]
    elif "/view" in base_url:
        base_url = base_url.split("/view")[0]

    return base_url.rstrip("/") + "/export?format=xlsx"


def normalize_cnpj(cnpj: str) -> str:
    return "".join(ch for ch in cnpj if ch.isdigit())


def build_tenant_domain(base: str) -> str:
    base = base.strip().lower()
    base = re.sub(r"[^a-z0-9]+", "-", base).strip("-")
    if not base:
        base = "tenant"
    return f"{base}.local"


def build_business_unit_code(name: str) -> str:
    if "matriz" in name.lower():
        return "MAT"
    cleaned = re.sub(r"[^a-zA-Z0-9]", "", name).upper()
    if len(cleaned) >= 3:
        return cleaned[:3]
    return (cleaned + "BU")[:3]

def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

@router.post("/clear-data")
async def clear_data(
    request: ClearDataRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Limpa dados financeiros de uma empresa/BU (mantém estrutura de contas se necessário)
    """
    try:
        # Normalizar strings vazias vindas do frontend
        if isinstance(request.tenant_id, str) and request.tenant_id.strip().lower() in {"", "undefined", "null"}:
            request.tenant_id = None
        if isinstance(request.business_unit_id, str) and request.business_unit_id.strip().lower() in {"", "undefined", "null"}:
            request.business_unit_id = None

        # Verificar se tenant e BU existem (quando informados)
        if request.tenant_id:
            tenant = db.query(Tenant).filter(
                Tenant.id == request.tenant_id,
                Tenant.status == "active",
            ).first()
            if not tenant:
                raise HTTPException(status_code=404, detail="Tenant não encontrado")

            if request.business_unit_id:
                business_unit = db.query(BusinessUnit).filter(
                    BusinessUnit.id == request.business_unit_id,
                    BusinessUnit.tenant_id == request.tenant_id,
                    BusinessUnit.status == "active",
                ).first()
                if not business_unit:
                    raise HTTPException(status_code=404, detail="Business Unit não encontrada")
        
        from app.models.lancamento_diario import LancamentoDiario
        from app.models.lancamento_previsto import LancamentoPrevisto
        from datetime import date
        
        deleted_counts = {
            "lancamentos_diarios": 0,
            "lancamentos_previstos": 0
        }
        
        # Deletar lançamentos diários
        if request.year:
            start_date = date(request.year, 1, 1)
            end_date = date(request.year, 12, 31)
            deleted_diarios = db.query(LancamentoDiario).filter(
                and_(
                    LancamentoDiario.tenant_id == request.tenant_id,
                    LancamentoDiario.business_unit_id == request.business_unit_id,
                    LancamentoDiario.data_movimentacao >= start_date,
                    LancamentoDiario.data_movimentacao <= end_date
                )
            ).delete(synchronize_session=False)
            deleted_counts["lancamentos_diarios"] = deleted_diarios
            
            deleted_previstos = db.query(LancamentoPrevisto).filter(
                and_(
                    LancamentoPrevisto.tenant_id == request.tenant_id,
                    LancamentoPrevisto.business_unit_id == request.business_unit_id,
                    LancamentoPrevisto.data_prevista >= start_date,
                    LancamentoPrevisto.data_prevista <= end_date
                )
            ).delete(synchronize_session=False)
            deleted_counts["lancamentos_previstos"] = deleted_previstos
        else:
            # Deletar todos os lançamentos
            deleted_diarios = db.query(LancamentoDiario).filter(
                and_(
                    LancamentoDiario.tenant_id == request.tenant_id,
                    LancamentoDiario.business_unit_id == request.business_unit_id
                )
            ).delete(synchronize_session=False)
            deleted_counts["lancamentos_diarios"] = deleted_diarios
            
            deleted_previstos = db.query(LancamentoPrevisto).filter(
                and_(
                    LancamentoPrevisto.tenant_id == request.tenant_id,
                    LancamentoPrevisto.business_unit_id == request.business_unit_id
                )
            ).delete(synchronize_session=False)
            deleted_counts["lancamentos_previstos"] = deleted_previstos
        
        db.commit()
        
        return {
            "success": True,
            "message": f"Dados limpos com sucesso",
            "deleted": deleted_counts,
            "year": request.year
        }
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao limpar dados: {str(e)}")

@router.post("/validate-spreadsheet")
async def validate_spreadsheet(
    request: SpreadsheetUrlRequest,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Valida se a URL da planilha é acessível e contém as abas necessárias
    """
    try:
        # Normalizar strings vazias vindas do frontend
        if isinstance(request.tenant_id, str) and request.tenant_id.strip().lower() in {"", "undefined", "null"}:
            request.tenant_id = None
        if isinstance(request.business_unit_id, str) and request.business_unit_id.strip().lower() in {"", "undefined", "null"}:
            request.business_unit_id = None

        # Verificar se tenant e BU existem (quando informados)
        if request.tenant_id:
            tenant = db.query(Tenant).filter(
                Tenant.id == request.tenant_id,
                Tenant.status == "active",
            ).first()
            if not tenant:
                raise HTTPException(status_code=404, detail="Tenant não encontrado")

            if request.business_unit_id:
                business_unit = db.query(BusinessUnit).filter(
                    BusinessUnit.id == request.business_unit_id,
                    BusinessUnit.tenant_id == request.tenant_id,
                    BusinessUnit.status == "active",
                ).first()
                if not business_unit:
                    raise HTTPException(status_code=404, detail="Business Unit não encontrada")
        
        # Baixar planilha temporariamente
        try:
            spreadsheet_url = normalize_spreadsheet_url(str(request.url))
            
            response = requests.get(spreadsheet_url, timeout=30)
            response.raise_for_status()
            
            # Salvar temporariamente
            with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
                tmp_file.write(response.content)
                tmp_file_path = tmp_file.name
            
            # Validar estrutura da planilha + consistência básica
            import pandas as pd
            from collections import defaultdict
            
            try:
                excel_file = pd.ExcelFile(tmp_file_path)
                required_sheets = ["Plano de contas", "Lançamento Diário", "Lançamentos Previstos"]
                available_sheets = excel_file.sheet_names
                
                # Verificar variações de nomes
                found_sheets = {}
                for req_sheet in required_sheets:
                    for avail_sheet in available_sheets:
                        if req_sheet.lower() in avail_sheet.lower() or avail_sheet.lower() in req_sheet.lower():
                            found_sheets[req_sheet] = avail_sheet
                            break
                
                missing_sheets = [s for s in required_sheets if s not in found_sheets]
                
                # Limpar arquivo temporário
                os.unlink(tmp_file_path)
                
                issues = []
                summary = {
                    "missing_sheets": missing_sheets,
                    "errors": 0,
                    "warnings": 0,
                }

                def add_issue(severity: str, code: str, message: str, sheet: str = None, row: int = None, data: dict = None):
                    issue_id = f"{code}:{sheet or 'sheet'}:{row or 'na'}"
                    issues.append(
                        {
                            "id": issue_id,
                            "severity": severity,
                            "code": code,
                            "message": message,
                            "sheet": sheet,
                            "row": row,
                            "data": data or {},
                        }
                    )
                    if severity == "error":
                        summary["errors"] += 1
                    elif severity == "warning":
                        summary["warnings"] += 1

                if missing_sheets:
                    return JSONResponse(
                        status_code=400,
                        content={
                            "valid": False,
                            "error": f"Abas faltantes: {', '.join(missing_sheets)}",
                            "available_sheets": available_sheets,
                            "found_sheets": found_sheets
                        }
                    )

                # ======== Carregar abas ========
                plano_df = pd.read_excel(excel_file, found_sheets["Plano de contas"])
                diarios_df = pd.read_excel(excel_file, found_sheets["Lançamento Diário"])
                previstos_df = pd.read_excel(excel_file, found_sheets["Lançamentos Previstos"])

                # ======== Helpers de colunas ========
                def _find_col(cols, includes):
                    for col in cols:
                        c = str(col).lower().strip()
                        if all(x in c for x in includes):
                            return col
                    return None

                # Plano de contas
                plano_cols = list(plano_df.columns)
                plano_conta_col = _find_col(plano_cols, ["conta"])
                plano_subgrupo_col = _find_col(plano_cols, ["subgrupo"])
                plano_grupo_col = _find_col(plano_cols, ["grupo"])

                plano_contas = set()
                if plano_conta_col:
                    plano_contas = {
                        str(v).strip().lower()
                        for v in plano_df[plano_conta_col].dropna().tolist()
                        if str(v).strip()
                    }

                # Lançamento Diário
                diarios_cols = list(diarios_df.columns)
                di_data_col = _find_col(diarios_cols, ["data", "moviment"])
                di_grupo_col = _find_col(diarios_cols, ["grupo"])
                di_subgrupo_col = _find_col(diarios_cols, ["subgrupo"])
                di_valor_col = _find_col(diarios_cols, ["valor"])

                if not di_data_col or not di_valor_col or not di_grupo_col or not di_subgrupo_col:
                    add_issue(
                        "error",
                        "DIARIO_COLUMNS_MISSING",
                        "Colunas obrigatórias não encontradas em Lançamento Diário (data, grupo, subgrupo, valor).",
                        sheet=found_sheets["Lançamento Diário"],
                    )
                else:
                    for idx, row in diarios_df.iterrows():
                        row_num = idx + 2
                        data_val = row.get(di_data_col)
                        valor_val = row.get(di_valor_col)
                        grupo_val = row.get(di_grupo_col)
                        subgrupo_val = row.get(di_subgrupo_col)

                        if pd.isna(data_val):
                            add_issue("error", "DIARIO_DATA_MISSING", "Data movimentação ausente.", found_sheets["Lançamento Diário"], row_num)
                            continue
                        if pd.isna(valor_val):
                            add_issue("error", "DIARIO_VALOR_MISSING", "Valor ausente.", found_sheets["Lançamento Diário"], row_num)
                            continue
                        if pd.isna(grupo_val) or pd.isna(subgrupo_val):
                            add_issue("error", "DIARIO_GRUPO_SUBGRUPO_MISSING", "Grupo/Subgrupo ausentes.", found_sheets["Lançamento Diário"], row_num)
                            continue

                # Lançamentos Previstos
                prev_cols = list(previstos_df.columns)
                prev_data_col = _find_col(prev_cols, ["data"]) or _find_col(prev_cols, ["mês"]) or _find_col(prev_cols, ["mes"])
                prev_conta_col = _find_col(prev_cols, ["conta"])
                prev_grupo_col = _find_col(prev_cols, ["grupo"])
                prev_subgrupo_col = _find_col(prev_cols, ["subgrupo"])
                prev_valor_col = _find_col(prev_cols, ["valor"])

                # Heurística para coluna de data: priorizar coluna com dia ≠ 1
                date_candidates = []
                for col in prev_cols:
                    try:
                        series = pd.to_datetime(previstos_df[col], errors="coerce")
                    except Exception:
                        continue
                    valid_ratio = series.notna().mean()
                    if valid_ratio >= 0.6:
                        day_not_one_ratio = (series.dt.day != 1).mean()
                        date_candidates.append((day_not_one_ratio, col))
                if date_candidates:
                    date_candidates.sort(reverse=True, key=lambda x: x[0])
                    best_ratio, best_col = date_candidates[0]
                    if best_ratio >= 0.1:
                        prev_data_col = best_col

                if not prev_data_col or not prev_conta_col or not prev_valor_col:
                    add_issue(
                        "error",
                        "PREVISTO_COLUMNS_MISSING",
                        "Colunas obrigatórias não encontradas em Lançamentos Previstos (data, conta, valor).",
                        sheet=found_sheets["Lançamentos Previstos"],
                    )
                else:
                    for idx, row in previstos_df.iterrows():
                        row_num = idx + 2
                        data_val = row.get(prev_data_col)
                        valor_val = row.get(prev_valor_col)
                        conta_val = row.get(prev_conta_col)
                        grupo_val = row.get(prev_grupo_col) if prev_grupo_col else None
                        subgrupo_val = row.get(prev_subgrupo_col) if prev_subgrupo_col else None

                        if pd.isna(data_val):
                            add_issue("error", "PREVISTO_DATA_MISSING", "Data prevista ausente.", found_sheets["Lançamentos Previstos"], row_num)
                            continue
                        if pd.isna(valor_val):
                            add_issue("error", "PREVISTO_VALOR_MISSING", "Valor ausente.", found_sheets["Lançamentos Previstos"], row_num)
                            continue
                        if pd.isna(conta_val) or str(conta_val).strip() == "":
                            add_issue("error", "PREVISTO_CONTA_MISSING", "Conta ausente.", found_sheets["Lançamentos Previstos"], row_num)
                            continue
                        if prev_grupo_col and prev_subgrupo_col and (pd.isna(grupo_val) or pd.isna(subgrupo_val)):
                            add_issue("error", "PREVISTO_GRUPO_SUBGRUPO_MISSING", "Grupo/Subgrupo ausentes.", found_sheets["Lançamentos Previstos"], row_num)
                            continue
                        if plano_contas and str(conta_val).strip().lower() not in plano_contas:
                            add_issue(
                                "error",
                                "PREVISTO_CONTA_FORA_PLANO",
                                f"Conta '{conta_val}' não encontrada no Plano de Contas.",
                                found_sheets["Lançamentos Previstos"],
                                row_num,
                            )

                # ======== Consistência Mensal (se existir Fluxo de caixa-2025) ========
                fluxo_sheet = None
                previsao_sheet = None
                for sheet in available_sheets:
                    sheet_lower = sheet.lower()
                    if "fluxo de caixa" in sheet_lower and "2025" in sheet_lower and "previs" not in sheet_lower:
                        fluxo_sheet = sheet
                        break
                if not fluxo_sheet:
                    for sheet in available_sheets:
                        sheet_lower = sheet.lower()
                        if "fluxo de caixa" in sheet_lower and "2025" in sheet_lower:
                            fluxo_sheet = sheet
                            break
                for sheet in available_sheets:
                    sheet_lower = sheet.lower()
                    if "fluxo de caixa" in sheet_lower and "2025" in sheet_lower and "previs" in sheet_lower:
                        previsao_sheet = sheet
                        break
                if fluxo_sheet:
                    fluxo_df = pd.read_excel(excel_file, fluxo_sheet, header=None)
                    # Encontrar linha de meses e linha de labels (Previsto/Realizado)
                    months_row = None
                    labels_row = None
                    for idx in range(min(10, len(fluxo_df))):
                        row = fluxo_df.iloc[idx]
                        if any(isinstance(val, str) and val.strip().upper() in MONTH_LABELS for val in row.values):
                            months_row = row
                            if idx + 1 < len(fluxo_df):
                                labels_row = fluxo_df.iloc[idx + 1]
                            break
                    if months_row is None:
                        months_row = fluxo_df.iloc[2]
                    if labels_row is None:
                        labels_row = fluxo_df.iloc[3] if len(fluxo_df) > 3 else fluxo_df.iloc[2]
                    month_cols = {}
                    for idx, val in months_row.items():
                        if isinstance(val, str) and val.strip():
                            month = val.strip().upper()
                            if str(labels_row.get(idx, "")) == "Previsto":
                                month_cols[month] = (idx, idx + 1)

                    label_col = 1
                    for col in fluxo_df.columns:
                        col_series = fluxo_df[col].astype(str).str.strip().str.lower()
                        if (col_series == "receita").any():
                            label_col = col
                            break

                    import unicodedata

                    def _norm(value: Any) -> str:
                        text = str(value or "")
                        normalized = unicodedata.normalize("NFKD", text)
                        normalized = "".join(ch for ch in normalized if not unicodedata.combining(ch))
                        return normalized.strip().lower()

                    label_row_index = {}
                    label_series = fluxo_df[label_col].astype(str).map(_norm)
                    for idx, val in label_series.items():
                        if val:
                            label_row_index[val] = idx
                    def get_row_value(label: str, month: str, kind: str) -> float:
                        row = fluxo_df[fluxo_df[label_col].astype(str).str.strip().str.lower() == label.strip().lower()]
                        if row.empty:
                            return 0.0
                        row = row.iloc[0]
                        col_prev, col_real = month_cols.get(month, (None, None))
                        val = row[col_real] if kind == "realizado" else row[col_prev]
                        if pd.isna(val):
                            return 0.0
                        return float(val)

                    # Agregar diários por grupo
                    if di_data_col and di_grupo_col and di_valor_col:
                        diarios_df["_data"] = pd.to_datetime(diarios_df[di_data_col], errors="coerce")
                        diarios_df["_mes"] = diarios_df["_data"].dt.month
                        diarios_df["_grupo"] = diarios_df[di_grupo_col].astype(str).map(_norm)
                        diarios_df["_valor"] = pd.to_numeric(diarios_df[di_valor_col], errors="coerce").fillna(0.0)

                        group_month_totals = defaultdict(lambda: defaultdict(float))
                        subgroup_month_totals = defaultdict(lambda: defaultdict(float))
                        diario_groups = set()
                        for _, row in diarios_df.iterrows():
                            if pd.isna(row.get("_data")):
                                continue
                            month = row["_mes"]
                            group = row["_grupo"]
                            if group:
                                diario_groups.add(group)
                            group_month_totals[group][month] += float(row["_valor"])
                            subgrupo_val = ""
                            if di_subgrupo_col:
                                subgrupo_val = _norm(row.get(di_subgrupo_col))
                            if group == _norm("movimentações não operacionais") and subgrupo_val:
                                subgroup_month_totals[subgrupo_val][month] += float(row["_valor"])

                        labels_to_check = [
                            "Receita",
                            "Deduções",
                            "Custos",
                            "Despesas Operacionais",
                            "Investimentos",
                            "Movimentações Não Operacionais",
                            "Entradas não Operacionais",
                            "Saídas não Operacionais",
                        ]
                        month_label_map = {i + 1: m for i, m in enumerate(MONTH_LABELS)}

                        child_labels_map = {
                            _norm("entradas não operacionais"): [
                                "Empréstimos/Financiamentos obtidos",
                                "Aporte dos sócios",
                                "Venda de equipamentos usados",
                                "Outros entradas não operacionais",
                            ],
                            _norm("saídas não operacionais"): [
                                "Giro Pronampe",
                                "Juros Bancários e por Atraso",
                                "Pagamento de Empréstimos",
                                "Retirada de Lucros",
                                "Juros de Antecipação de Recebíveis",
                                "Outras saídas não operacionais",
                            ],
                        }

                        for label in labels_to_check:
                            label_key = _norm(label)
                            if label_key in {"entradas não operacionais", "saídas não operacionais"}:
                                # Se existirem linhas filhas, comparar com elas e ignorar o totalizador
                                child_labels = child_labels_map.get(label_key, [])
                                has_child_rows = False
                                for child in child_labels:
                                    child_key = _norm(child)
                                    if child_key in label_row_index:
                                        has_child_rows = True
                                        for month_num, month_name in month_label_map.items():
                                            sheet_real = get_row_value(child, month_name, "realizado")
                                            diario_real = subgroup_month_totals.get(child_key, {}).get(month_num, 0.0)
                                            diff = round(diario_real - sheet_real, 2)
                                            if abs(diff) > 0.0:
                                                row_idx = label_row_index.get(child_key)
                                                row_ref = (row_idx + 1) if row_idx is not None else None
                                                add_issue(
                                                    "error",
                                                    "MENSAL_DIARIO_DIVERGENCE",
                                                    f"'{child}' {month_name}: planilha={sheet_real:.2f} vs diário={diario_real:.2f} (dif={diff:.2f})",
                                                    sheet=fluxo_sheet,
                                                    row=row_ref,
                                                    data={"label": child, "month": month_name},
                                                )
                                if has_child_rows:
                                    continue
                                # Se a linha da planilha não possui valores preenchidos, não comparar
                                total_sheet = 0.0
                                for month_name in month_label_map.values():
                                    total_sheet += abs(get_row_value(label, month_name, "realizado"))
                                if total_sheet == 0.0:
                                    continue
                                # Comparar pelo subgrupo quando grupo for Movimentações Não Operacionais
                                for month_num, month_name in month_label_map.items():
                                    sheet_real = get_row_value(label, month_name, "realizado")
                                    diario_real = subgroup_month_totals.get(label_key, {}).get(month_num, 0.0)
                                    diff = round(diario_real - sheet_real, 2)
                                    if abs(diff) > 0.0:
                                        row_idx = label_row_index.get(label_key)
                                        row_ref = (row_idx + 1) if row_idx is not None else None
                                        add_issue(
                                            "error",
                                            "MENSAL_DIARIO_DIVERGENCE",
                                            f"'{label}' {month_name}: planilha={sheet_real:.2f} vs diário={diario_real:.2f} (dif={diff:.2f})",
                                            sheet=fluxo_sheet,
                                            row=row_ref,
                                            data={"label": label, "month": month_name},
                                        )
                                continue

                            if label_key not in diario_groups:
                                # Não comparar linhas totalizadoras que não existem no diário
                                continue
                            for month_num, month_name in month_label_map.items():
                                sheet_real = get_row_value(label, month_name, "realizado")
                                diario_real = group_month_totals[label_key].get(month_num, 0.0)
                                diff = round(diario_real - sheet_real, 2)
                                if abs(diff) > 0.0:
                                    row_idx = label_row_index.get(label_key)
                                    row_ref = (row_idx + 1) if row_idx is not None else None
                                    add_issue(
                                        "error",
                                        "MENSAL_DIARIO_DIVERGENCE",
                                        f"'{label}' {month_name}: planilha={sheet_real:.2f} vs diário={diario_real:.2f} (dif={diff:.2f})",
                                        sheet=fluxo_sheet,
                                        row=row_ref,
                                        data={"label": label, "month": month_name},
                                    )
                if previsao_sheet:
                    previsao_df = pd.read_excel(excel_file, previsao_sheet, header=None)
                    previsao_months_row = None
                    for idx in range(min(10, len(previsao_df))):
                        row = previsao_df.iloc[idx]
                        if any(isinstance(val, str) and val.strip().upper() in MONTH_LABELS for val in row.values):
                            previsao_months_row = row
                            break
                    if previsao_months_row is None:
                        add_issue(
                            "error",
                            "FORECAST_SHEET_INVALID",
                            "Não foi possível identificar os meses na aba de Previsão Fluxo de caixa-2025.",
                            sheet=previsao_sheet,
                        )

                # Resultado final
                summary.update(
                    {
                        "available_sheets": available_sheets,
                        "found_sheets": found_sheets,
                    }
                )

                return {
                    "valid": summary["errors"] == 0,
                    "message": "Planilha validada com relatório de inconsistências",
                    "available_sheets": available_sheets,
                    "found_sheets": found_sheets,
                    "summary": summary,
                    "issues": issues,
                }
                
            except Exception as e:
                if os.path.exists(tmp_file_path):
                    os.unlink(tmp_file_path)
                raise HTTPException(status_code=400, detail=f"Erro ao validar planilha: {str(e)}")
                
        except requests.RequestException as e:
            raise HTTPException(status_code=400, detail=f"Erro ao acessar planilha: {str(e)}")
            
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao validar planilha: {str(e)}")

@router.post("/start")
async def start_onboarding(
    request: StartOnboardingRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(require_super_admin),
    db: Session = Depends(get_db),
):
    """
    Cria tenant + BU e inicia a importação via planilha.
    """
    try:
        normalized_cnpj = normalize_cnpj(request.tenant_cnpj)
        if not normalized_cnpj:
            raise HTTPException(status_code=400, detail="CNPJ inválido")

        tenant = db.query(Tenant).filter(
            Tenant.cnpj == normalized_cnpj,
            Tenant.status != "deleted",
        ).first()
        normalized_spreadsheet_url = normalize_spreadsheet_url(str(request.spreadsheet_url))
        if not tenant:
            base_domain = f"cnpj-{normalized_cnpj}"
            domain = build_tenant_domain(base_domain)
            if db.query(Tenant).filter(Tenant.domain == domain).first():
                domain = build_tenant_domain(f"{base_domain}-{str(uuid.uuid4())[:8]}")

            tenant = Tenant(
                name=request.tenant_name.strip(),
                domain=domain,
                cnpj=normalized_cnpj,
                spreadsheet_url=normalized_spreadsheet_url,
                status="active",
            )
            db.add(tenant)
            db.commit()
            db.refresh(tenant)
        elif normalized_spreadsheet_url and tenant.spreadsheet_url != normalized_spreadsheet_url:
            tenant.spreadsheet_url = normalized_spreadsheet_url
            tenant.status = "active"
            db.commit()

        business_unit = (
            db.query(BusinessUnit)
            .filter(
                BusinessUnit.tenant_id == tenant.id,
                BusinessUnit.name == request.business_unit_name.strip(),
                BusinessUnit.status == "active",
            )
            .first()
        )

        if not business_unit:
            bu_code = request.business_unit_code or build_business_unit_code(request.business_unit_name)
            business_unit = BusinessUnit(
                tenant_id=tenant.id,
                name=request.business_unit_name.strip(),
                code=bu_code,
                status="active",
            )
            db.add(business_unit)
            db.commit()
            db.refresh(business_unit)

        status_key = f"{tenant.id}_{business_unit.id}"
        run_id = str(uuid.uuid4())
        onboarding_status[status_key] = init_status(
            str(tenant.id),
            str(business_unit.id),
            "Iniciando validação da planilha...",
            run_id,
        )
        log_onboarding(
            "onboarding_started",
            status_key=status_key,
            run_id=run_id,
            tenant_id=str(tenant.id),
            business_unit_id=str(business_unit.id),
            spreadsheet_url=normalized_spreadsheet_url,
            reset_data=request.reset_data,
            user_id=str(current_user.id),
        )

        background_tasks.add_task(
            execute_import,
            str(tenant.id),
            str(business_unit.id),
            normalized_spreadsheet_url,
            request.reset_data,
            str(current_user.id),
            request.corrections,
        )

        return {
            "success": True,
            "message": "Onboarding iniciado",
            "tenant_id": str(tenant.id),
            "business_unit_id": str(business_unit.id),
            "status_key": status_key,
            "run_id": run_id,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"Erro ao iniciar onboarding: {str(e)}")

@router.post("/import")
async def import_data(
    request: ImportRequest,
    background_tasks: BackgroundTasks,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Inicia processo de importação de dados em etapas
    """
    try:
        # Verificar se tenant e BU existem
        tenant = db.query(Tenant).filter(
            Tenant.id == request.tenant_id,
            Tenant.status == "active",
        ).first()
        if not tenant:
            raise HTTPException(status_code=404, detail="Tenant não encontrado")
        
        business_unit = db.query(BusinessUnit).filter(
            BusinessUnit.id == request.business_unit_id,
            BusinessUnit.tenant_id == request.tenant_id,
            BusinessUnit.status == "active",
        ).first()
        if not business_unit:
            raise HTTPException(status_code=404, detail="Business Unit não encontrada")
        
        corrections = request.corrections or {}
        row_updates = corrections.get("row_updates") or []
        ignore_issue_ids = corrections.get("ignore_issue_ids") or []
        if (row_updates or ignore_issue_ids) and corrections.get("confirm") is not True:
            raise HTTPException(status_code=400, detail="Confirmação obrigatória para aplicar correções manuais.")

        if row_updates or ignore_issue_ids:
            db.add(
                AuditLog(
                    user_id=str(current_user.id),
                    tenant_id=str(request.tenant_id),
                    action="onboarding_corrections",
                    resource_type="onboarding",
                    resource_id=f"{request.tenant_id}:{request.business_unit_id}",
                    details=json.dumps(
                        {
                            "spreadsheet_url": str(request.spreadsheet_url),
                            "corrections": corrections,
                        },
                        ensure_ascii=True,
                    ),
                )
            )
            db.commit()

        status_key = f"{request.tenant_id}_{request.business_unit_id}"
        
        # Inicializar status
        run_id = str(uuid.uuid4())
        onboarding_status[status_key] = init_status(
            request.tenant_id,
            request.business_unit_id,
            "Iniciando validação da planilha...",
            run_id,
        )
        normalized_spreadsheet_url = normalize_spreadsheet_url(str(request.spreadsheet_url))
        if tenant.spreadsheet_url != normalized_spreadsheet_url:
            tenant.spreadsheet_url = normalized_spreadsheet_url
            db.commit()

        log_onboarding(
            "import_started",
            status_key=status_key,
            run_id=run_id,
            tenant_id=request.tenant_id,
            business_unit_id=request.business_unit_id,
            spreadsheet_url=normalized_spreadsheet_url,
            reset_data=request.reset_data,
            user_id=str(current_user.id),
        )
        
        # Executar importação em background
        background_tasks.add_task(
            execute_import,
            request.tenant_id,
            request.business_unit_id,
            normalized_spreadsheet_url,
            request.reset_data,
            current_user.id,
            corrections,
        )
        
        return {
            "success": True,
            "message": "Importação iniciada",
            "status_key": status_key,
            "run_id": run_id
        }
        
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao iniciar importação: {str(e)}")

@router.get("/status/{tenant_id}/{business_unit_id}")
async def get_onboarding_status(
    tenant_id: str,
    business_unit_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db),
):
    """
    Retorna status atual do onboarding
    """
    business_unit = db.query(BusinessUnit).join(Tenant, BusinessUnit.tenant_id == Tenant.id).filter(
        Tenant.id == tenant_id,
        Tenant.status == "active",
        BusinessUnit.id == business_unit_id,
        BusinessUnit.status == "active",
    ).first()
    if not business_unit:
        raise HTTPException(status_code=404, detail="Tenant ou Business Unit não encontrada")

    status_key = f"{tenant_id}_{business_unit_id}"
    status = onboarding_status.get(status_key)
    
    if not status:
        return {
            "status": "not_started",
            "progress": 0,
            "message": "Onboarding não iniciado"
        }
    
    return status.dict()

@router.get("/reconciliation/{tenant_id}/{business_unit_id}")
async def get_reconciliation(
    tenant_id: str,
    business_unit_id: str,
    current_user: User = Depends(get_current_active_user),
    db: Session = Depends(get_db)
):
    """
    Retorna relatório de conciliação entre planilha e sistema
    """
    try:
        business_unit = db.query(BusinessUnit).join(Tenant, BusinessUnit.tenant_id == Tenant.id).filter(
            Tenant.id == tenant_id,
            Tenant.status == "active",
            BusinessUnit.id == business_unit_id,
            BusinessUnit.status == "active",
        ).first()
        if not business_unit:
            raise HTTPException(status_code=404, detail="Tenant ou Business Unit não encontrada")

        # Buscar última planilha importada para este tenant/BU
        data_dir = backend_path / "data"
        excel_files = sorted(
            [f for f in data_dir.glob(f"onboarding_{tenant_id}_{business_unit_id}_*.xlsx")],
            key=lambda x: x.stat().st_mtime,
            reverse=True
        )
        
        if not excel_files:
            return {
                "tenant_id": tenant_id,
                "business_unit_id": business_unit_id,
                "reconciliation_date": datetime.now().isoformat(),
                "status": "no_data",
                "message": "Nenhuma planilha importada encontrada"
            }
        
        excel_file_path = excel_files[0]
        
        # Importar funções de conciliação
        from scripts.reconcile_fluxo_caixa import extract_fluxo_caixa_totals
        import requests
        import os
        
        # Extrair totais da planilha
        excel_totals_dict = extract_fluxo_caixa_totals(excel_file_path)
        
        # Converter para formato esperado
        excel_totals = {
            "receita_total": sum(float(excel_totals_dict.get(m, {}).get("receita", 0)) for m in range(1, 13)),
            "despesa_total": sum(float(excel_totals_dict.get(m, {}).get("despesa", 0)) for m in range(1, 13)),
            "custo_total": sum(float(excel_totals_dict.get(m, {}).get("custo", 0)) for m in range(1, 13)),
            "saldo_total": sum(float(excel_totals_dict.get(m, {}).get("saldo", 0)) for m in range(1, 13)),
            "mensal": {
                str(m): {
                    "receita": float(excel_totals_dict.get(m, {}).get("receita", 0)),
                    "despesa": float(excel_totals_dict.get(m, {}).get("despesa", 0)),
                    "custo": float(excel_totals_dict.get(m, {}).get("custo", 0)),
                    "saldo": float(excel_totals_dict.get(m, {}).get("saldo", 0))
                }
                for m in range(1, 13)
            }
        }
        
        # Buscar totais da API
        BACKEND_URL = os.getenv("BACKEND_URL", "https://finaflow-backend-staging-556803510516.us-central1.run.app")
        QA_USERNAME = os.getenv("QA_USERNAME", "qa@finaflow.test")
        QA_PASSWORD = os.getenv("QA_PASSWORD", "QaFinaflow123!")
        
        login_resp = requests.post(
            f"{BACKEND_URL}/api/v1/auth/login",
            json={"username": QA_USERNAME, "password": QA_PASSWORD},
            timeout=30
        )
        token = login_resp.json()["access_token"]
        headers = {"Authorization": f"Bearer {token}"}
        
        api_resp = requests.get(
            f"{BACKEND_URL}/api/v1/financial/annual-summary?year=2025",
            headers=headers,
            timeout=30
        )
        api_data = api_resp.json()
        
        # Calcular diferenças
        totals = {
            "revenue": {
                "excel": float(excel_totals.get("receita_total", 0)),
                "system": float(api_data.get("totals", {}).get("revenue", 0)),
                "diff": float(excel_totals.get("receita_total", 0)) - float(api_data.get("totals", {}).get("revenue", 0))
            },
            "expense": {
                "excel": float(excel_totals.get("despesa_total", 0)),
                "system": float(api_data.get("totals", {}).get("expense", 0)),
                "diff": float(excel_totals.get("despesa_total", 0)) - float(api_data.get("totals", {}).get("expense", 0))
            },
            "cost": {
                "excel": float(excel_totals.get("custo_total", 0)),
                "system": float(api_data.get("totals", {}).get("cost", 0)),
                "diff": float(excel_totals.get("custo_total", 0)) - float(api_data.get("totals", {}).get("cost", 0))
            },
            "balance": {
                "excel": float(excel_totals.get("saldo_total", 0)),
                "system": float(api_data.get("totals", {}).get("balance", 0)),
                "diff": float(excel_totals.get("saldo_total", 0)) - float(api_data.get("totals", {}).get("balance", 0))
            }
        }
        
        # Calcular mensais
        monthly = []
        for month in range(1, 13):
            month_data: Dict[str, Any] = {"month": month}
            
            excel_month = excel_totals.get("mensal", {}).get(str(month), {})
            api_month = api_data.get("monthly", {}).get(str(month), {})
            
            if excel_month.get("receita") is not None:
                month_data["revenue"] = {
                    "excel": float(excel_month.get("receita", 0)),
                    "system": float(api_month.get("revenue", 0)),
                    "diff": float(excel_month.get("receita", 0)) - float(api_month.get("revenue", 0))
                }
            
            if excel_month.get("despesa") is not None:
                month_data["expense"] = {
                    "excel": float(excel_month.get("despesa", 0)),
                    "system": float(api_month.get("expense", 0)),
                    "diff": float(excel_month.get("despesa", 0)) - float(api_month.get("expense", 0))
                }
            
            if excel_month.get("custo") is not None:
                month_data["cost"] = {
                    "excel": float(excel_month.get("custo", 0)),
                    "system": float(api_month.get("cost", 0)),
                    "diff": float(excel_month.get("custo", 0)) - float(api_month.get("cost", 0))
                }
            
            if excel_month.get("saldo") is not None:
                month_data["balance"] = {
                    "excel": float(excel_month.get("saldo", 0)),
                    "system": float(api_month.get("balance", 0)),
                    "diff": float(excel_month.get("saldo", 0)) - float(api_month.get("balance", 0))
                }
            
            monthly.append(month_data)
        
        # Verificar se está 100% conciliado
        all_reconciled = all(
            abs(t["diff"]) < 0.01
            for t in totals.values()
        ) and all(
            all(abs(v.get("diff", 0)) < 0.01 for v in m.values() if isinstance(v, dict) and "diff" in v)
            for m in monthly
        )
        
        return {
            "tenant_id": tenant_id,
            "business_unit_id": business_unit_id,
            "reconciliation_date": datetime.now().isoformat(),
            "status": "completed" if all_reconciled else "differences_found",
            "totals": totals,
            "monthly": monthly,
            "all_reconciled": all_reconciled
        }
        
    except Exception as e:
        import traceback
        raise HTTPException(
            status_code=500,
            detail=f"Erro ao gerar conciliação: {str(e)}\n{traceback.format_exc()}"
        )

def apply_corrections_to_excel(excel_file_path: Path, corrections: Optional[Dict[str, Any]]) -> Dict[str, int]:
    if not corrections:
        return {"applied": 0, "skipped": 0}

    row_updates = corrections.get("row_updates") or []
    if not row_updates:
        return {"applied": 0, "skipped": 0}

    try:
        from openpyxl import load_workbook
    except Exception:
        return {"applied": 0, "skipped": len(row_updates)}

    wb = load_workbook(excel_file_path)
    applied = 0
    skipped = 0

    for update in row_updates:
        sheet_name = update.get("sheet")
        row_num = update.get("row")
        fields = update.get("fields") or {}
        if not sheet_name or not row_num or not isinstance(fields, dict) or not fields:
            skipped += 1
            continue
        if sheet_name not in wb.sheetnames:
            skipped += 1
            continue

        ws = wb[sheet_name]
        headers = {}
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val is None:
                continue
            headers[str(header_val).strip().lower()] = col_idx

        def find_col(keys: List[str]) -> Optional[int]:
            for key in keys:
                for header, col_idx in headers.items():
                    if key in header:
                        return col_idx
            return None

        col_map = {
            "data": find_col(["data"]) or find_col(["mês"]) or find_col(["mes"]),
            "valor": find_col(["valor"]),
            "grupo": find_col(["grupo"]),
            "subgrupo": find_col(["subgrupo"]),
            "conta": find_col(["conta"]),
        }

        updated = False
        for field, value in fields.items():
            col_idx = col_map.get(field)
            if not col_idx:
                continue
            ws.cell(row=row_num, column=col_idx, value=value)
            updated = True
        if updated:
            applied += 1
        else:
            skipped += 1

    wb.save(excel_file_path)
    return {"applied": applied, "skipped": skipped}


def execute_import(
    tenant_id: str,
    business_unit_id: str,
    spreadsheet_url: str,
    reset_data: bool,
    user_id: str,
    corrections: Optional[Dict[str, Any]] = None,
):
    """
    Executa importação em etapas (background task)
    """
    status_key = f"{tenant_id}_{business_unit_id}"
    start_time = time.monotonic()
    
    try:
        # Etapa 1: Baixar planilha
        update_status(
            status_key,
            status="validating",
            current_step="Baixando planilha",
            progress=10,
            message="Baixando planilha da URL...",
        )
        
        spreadsheet_url = normalize_spreadsheet_url(spreadsheet_url)
        log_onboarding(
            "download_started",
            status_key=status_key,
            spreadsheet_url=spreadsheet_url,
            tenant_id=tenant_id,
            business_unit_id=business_unit_id,
            user_id=user_id,
        )
        
        response = requests.get(spreadsheet_url, timeout=60)
        response.raise_for_status()
        
        # Salvar planilha temporariamente
        data_dir = backend_path / "data"
        data_dir.mkdir(exist_ok=True)
        excel_file_path = data_dir / f"onboarding_{tenant_id}_{business_unit_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        with open(excel_file_path, "wb") as f:
            f.write(response.content)
        log_onboarding(
            "download_completed",
            status_key=status_key,
            file_path=str(excel_file_path),
            file_size=len(response.content),
        )
        
        # Aplicar correções manuais (se houver)
        correction_stats = apply_corrections_to_excel(Path(excel_file_path), corrections)
        if correction_stats.get("applied") or correction_stats.get("skipped"):
            log_onboarding(
                "corrections_applied",
                status_key=status_key,
                applied=correction_stats.get("applied"),
                skipped=correction_stats.get("skipped"),
            )

        # Etapa 2: Importar Plano de Contas
        update_status(
            status_key,
            status="importing_plan",
            current_step="Importando Plano de Contas",
            progress=30,
            message="Importando estrutura de contas...",
        )
        
        # Executar seed do plano de contas
        from scripts.seed_from_client_sheet import seed_plano_contas
        from app.database import SessionLocal
        from app.models.auth import Tenant, BusinessUnit
        from app.models.lancamento_diario import LancamentoDiario
        from app.models.lancamento_previsto import LancamentoPrevisto
        
        db = SessionLocal()
        try:
            # Buscar tenant, BU e user existentes
            tenant = db.query(Tenant).filter(
                Tenant.id == tenant_id,
                Tenant.status == "active",
            ).first()
            if not tenant:
                raise Exception(f"Tenant {tenant_id} não encontrado")
            
            business_unit = db.query(BusinessUnit).filter(
                BusinessUnit.id == business_unit_id,
                BusinessUnit.tenant_id == tenant_id,
                BusinessUnit.status == "active",
            ).first()
            if not business_unit:
                raise Exception(f"Business Unit {business_unit_id} não encontrada")
            
            grupos_map, subgrupos_map, contas_map = seed_plano_contas(db, tenant, excel_file_path)
            
            update_status(
                status_key,
                progress=50,
                message=f"Plano de contas importado: {len(contas_map)} contas",
                stats={
                    "contas_importadas": len(contas_map),
                    "grupos_importados": len(grupos_map),
                    "subgrupos_importados": len(subgrupos_map),
                },
            )
            
            # Buscar ou criar usuário para o seed
            if str(backend_path) not in sys.path:
                sys.path.insert(0, str(backend_path))
            from scripts.seed_from_client_sheet import get_or_create_user
            user = get_or_create_user(db, tenant, business_unit, user_id)
            
            # Etapa 3: Importar Lançamentos
            update_status(
                status_key,
                status="importing_transactions",
                current_step="Importando Lançamentos Financeiros",
                progress=60,
                message="Importando lançamentos diários e previstos...",
            )
            
            # Executar seed diretamente (sem subprocess para evitar timeout)
            from scripts.seed_from_client_sheet import seed_lancamentos_previstos, seed_lancamentos_diarios
            
            # Resetar dados se solicitado
            if reset_data:
                from datetime import date
                from sqlalchemy import and_
                
                deleted_diarios = db.query(LancamentoDiario).filter(
                    and_(
                        LancamentoDiario.tenant_id == tenant.id,
                        LancamentoDiario.data_movimentacao >= date(2025, 1, 1),
                        LancamentoDiario.data_movimentacao <= date(2025, 12, 31)
                    )
                ).delete(synchronize_session=False)
                
                deleted_prev = db.query(LancamentoPrevisto).filter(
                    and_(
                        LancamentoPrevisto.tenant_id == tenant.id,
                        LancamentoPrevisto.data_prevista >= date(2025, 1, 1),
                        LancamentoPrevisto.data_prevista <= date(2025, 12, 31)
                    )
                ).delete(synchronize_session=False)
                
                db.commit()
                update_status(
                    status_key,
                    message=f"Dados resetados: {deleted_diarios} diários, {deleted_prev} previstos",
                )
            
            # Importar lançamentos previstos
            update_status(
                status_key,
                progress=65,
                message="Importando lançamentos previstos...",
            )
            try:
                # Usar logger global do módulo seed
                from scripts.seed_from_client_sheet import logger as seed_logger
                previstos_before = seed_logger.stats.get('lancamentos_previstos_criados', 0)
                seed_lancamentos_previstos(db, tenant, business_unit, user, grupos_map, subgrupos_map, contas_map, Path(excel_file_path))
                db.commit()  # Commit após previstos
                previstos_after = seed_logger.stats.get('lancamentos_previstos_criados', 0)
                previstos_count = previstos_after - previstos_before
                update_status(
                    status_key,
                    message=f"Lançamentos previstos importados: {previstos_count}",
                )
            except Exception as e:
                db.rollback()
                import traceback
                error_trace = traceback.format_exc()
                raise Exception(f"Erro ao importar lançamentos previstos: {str(e)}\n{error_trace}")
            
            # Importar lançamentos diários
            update_status(
                status_key,
                progress=80,
                message="Importando lançamentos diários (pode levar alguns minutos)...",
            )
            try:
                # Usar logger global do módulo seed
                from scripts.seed_from_client_sheet import logger as seed_logger
                import threading
                import time as time_module
                
                # Resetar stats do logger para esta execução
                diarios_before = seed_logger.stats.get('lancamentos_diarios_criados', 0)
                previstos_before = seed_logger.stats.get('lancamentos_previstos_criados', 0)
                
                print(f"🔍 [ONBOARDING] Antes da importação: Diários={diarios_before}, Previstos={previstos_before}")
                
                # Função para atualizar progresso periodicamente durante importação
                def update_progress():
                    while True:
                        time_module.sleep(15)  # Atualizar a cada 15 segundos
                        current_diarios = seed_logger.stats.get('lancamentos_diarios_criados', 0)
                        current_previstos = seed_logger.stats.get('lancamentos_previstos_criados', 0)
                        current_created_diarios = current_diarios - diarios_before
                        current_created_previstos = current_previstos - previstos_before
                        update_status(
                            status_key,
                            message=f"Importando... Diários: {current_created_diarios} | Previstos: {current_created_previstos}",
                        )
                        print(f"📊 [ONBOARDING] Progresso: Diários={current_created_diarios}, Previstos={current_created_previstos}")
                        if onboarding_status[status_key].status == "completed" or onboarding_status[status_key].status == "error":
                            break
                
                # Iniciar thread de atualização de progresso
                progress_thread = threading.Thread(target=update_progress, daemon=True)
                progress_thread.start()
                
                # Executar importação com logging detalhado
                print(f"🚀 [ONBOARDING] Iniciando seed_lancamentos_diarios...")
                print(f"   Excel file: {excel_file_path}")
                print(f"   Tenant: {tenant.name} ({tenant.id})")
                print(f"   BU: {business_unit.name} ({business_unit.id})")
                print(f"   Grupos: {len(grupos_map)}, Subgrupos: {len(subgrupos_map)}, Contas: {len(contas_map)}")
                
                seed_lancamentos_diarios(db, tenant, business_unit, user, grupos_map, subgrupos_map, contas_map, Path(excel_file_path))
                
                print(f"✅ [ONBOARDING] seed_lancamentos_diarios concluído")
                print(f"   Stats após execução: {seed_logger.stats}")
                
                # Commit final (a função já faz commits em lotes, mas garantimos o commit final)
                try:
                    db.commit()
                    print(f"✅ [ONBOARDING] Commit final realizado")
                except Exception as commit_error:
                    print(f"⚠️ [ONBOARDING] Erro no commit final (pode ser que já foi commitado): {commit_error}")
                    db.rollback()
                    # Tentar novamente
                    try:
                        db.commit()
                        print(f"✅ [ONBOARDING] Commit final realizado na segunda tentativa")
                    except:
                        pass
                
                diarios_after = seed_logger.stats.get('lancamentos_diarios_criados', 0)
                previstos_after = seed_logger.stats.get('lancamentos_previstos_criados', 0)
                diarios_count = diarios_after - diarios_before
                previstos_count = previstos_after - previstos_before
                
                print(f"📊 [ONBOARDING] Resultado final: Diários criados={diarios_count}, Previstos criados={previstos_count}")
                
                update_status(
                    status_key,
                    progress=90,
                    message=f"Lançamentos importados! Diários: {diarios_count}, Previstos: {previstos_count}",
                    stats={
                        "contas_importadas": len(contas_map),
                        "grupos_importados": len(grupos_map),
                        "subgrupos_importados": len(subgrupos_map),
                        "lancamentos_diarios_criados": diarios_count,
                        "lancamentos_previstos_criados": previstos_count,
                    },
                )

                # Persistir configurações do fluxo de caixa (ordem e saldo inicial)
                from app.models.cash_flow_settings import CashFlowYearSettings
                from app.models.cash_flow_forecast_values import CashFlowForecastValue
                from scripts.seed_from_client_sheet import (
                    extract_cash_flow_settings,
                    extract_cash_flow_forecast_values,
                )

                settings_by_year = extract_cash_flow_settings(Path(excel_file_path))
                forecast_by_year = extract_cash_flow_forecast_values(Path(excel_file_path))
                for year_key, data in settings_by_year.items():
                    line_order = data.get("line_order") or []
                    saldo_ano_anterior = Decimal(str(data.get("saldo_ano_anterior") or 0))
                    existing = (
                        db.query(CashFlowYearSettings)
                        .filter(
                            CashFlowYearSettings.tenant_id == tenant.id,
                            CashFlowYearSettings.business_unit_id == business_unit.id,
                            CashFlowYearSettings.year == int(year_key),
                        )
                        .first()
                    )
                    if existing:
                        existing.saldo_ano_anterior = saldo_ano_anterior
                        existing.line_order = json.dumps(line_order, ensure_ascii=True)
                    else:
                        db.add(
                            CashFlowYearSettings(
                                tenant_id=str(tenant.id),
                                business_unit_id=str(business_unit.id),
                                year=int(year_key),
                                saldo_ano_anterior=saldo_ano_anterior,
                                line_order=json.dumps(line_order, ensure_ascii=True),
                            )
                        )

                # Atualizar valores previstos do fluxo de caixa
                for year_key, values in forecast_by_year.items():
                    for label, months in values.items():
                        for month_label, value in months.items():
                            month_index = None
                            try:
                                month_index = MONTH_LABELS.index(month_label) + 1
                            except Exception:
                                month_index = None
                            if not month_index:
                                continue
                            existing_value = (
                                db.query(CashFlowForecastValue)
                                .filter(
                                    CashFlowForecastValue.tenant_id == tenant.id,
                                    CashFlowForecastValue.business_unit_id == business_unit.id,
                                    CashFlowForecastValue.year == int(year_key),
                                    CashFlowForecastValue.label == label,
                                    CashFlowForecastValue.month == month_index,
                                )
                                .first()
                            )
                            if existing_value:
                                existing_value.value = Decimal(str(value or 0))
                            else:
                                db.add(
                                    CashFlowForecastValue(
                                        tenant_id=str(tenant.id),
                                        business_unit_id=str(business_unit.id),
                                        year=int(year_key),
                                        label=label,
                                        month=month_index,
                                        value=Decimal(str(value or 0)),
                                    )
                                )

                db.commit()
            except Exception as e:
                db.rollback()
                import traceback
                error_trace = traceback.format_exc()
                print(f"❌ [ONBOARDING] ERRO DETALHADO na importação de lançamentos diários:")
                print(error_trace)
                update_status(
                    status_key,
                    status="error",
                    message=f"Erro ao importar lançamentos diários: {str(e)}",
                    error=str(e),
                    finished=True,
                )
                update_status(status_key, error=error_trace)
                return
                
        except Exception as e:
            db.rollback()
            import traceback
            update_status(
                status_key,
                status="error",
                message=f"Erro ao importar lançamentos: {str(e)}",
                error=str(e),
                finished=True,
            )
            update_status(status_key, error=traceback.format_exc())
            return
        finally:
            db.close()
        
        # Etapa 4: Conciliação
        update_status(
            status_key,
            status="reconciling",
            current_step="Conciliação",
            progress=95,
            message="Gerando relatório de conciliação...",
        )
        
        # TODO: Executar conciliação
        
        # Finalizar
        duration_seconds = round(time.monotonic() - start_time, 2)
        update_status(
            status_key,
            status="completed",
            current_step="Concluído",
            progress=100,
            message="Onboarding concluído com sucesso!",
            stats={
                "contas_importadas": len(contas_map),
                "grupos_importados": len(grupos_map),
                "subgrupos_importados": len(subgrupos_map),
                "duration_seconds": duration_seconds,
            },
            finished=True,
        )
        log_onboarding(
            "onboarding_completed",
            status_key=status_key,
            tenant_id=tenant_id,
            business_unit_id=business_unit_id,
            duration_seconds=duration_seconds,
        )
        
    except Exception as e:
        update_status(
            status_key,
            status="error",
            message=f"Erro durante onboarding: {str(e)}",
            error=str(e),
            finished=True,
        )
        log_onboarding(
            "onboarding_failed",
            status_key=status_key,
            tenant_id=tenant_id,
            business_unit_id=business_unit_id,
            error=str(e),
        )
